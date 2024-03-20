# # # print('xxx',end='')
# # import random
# # import string
# # import unicodedata


# # type("x")
# # id("x")
# # random.randrange(0, 100)
# # print("\a")
# # len("xxxxx")
# # ":".join(["a", "c", "d"])
# # "大阪" in "私は大阪府に住んでいます．"
# # "文文字列".find("字")
# # "o".index("word")
# # "ABCdefghiJKL".startswith("ABC")
# # "ABCdefghiJKL".endswith("JKL")
# # "ABCdefghiJKL".count("JKL")
# # "ABCdefghiJKL".strip("JKL")
# # "   ABCdefghiJKL   ".strip()
# # ord("\r")
# # chr(13)
# # unicodedata.name("/r")
# # str(None)
# # type([1, 2, 3])
# # type([1, 2, 3]) is list
# # isinstance(3, int)
# # (3.0).is_integer()
# # l = []
# # [].append("z")
# # [].insert(1, "b")
# # del l[1]
# # l.remove[1]
# # l = [1, 2, 3, 4, 5]
# # 2 in l
# # l.index(3)

# # try:
# #     n = l.index(3)
# #     print(n)
# # except ValueError:
# #     print("cccc")

# # l.count(1)
# # len(l)
# # sum(l)
# # l.sort()

# # l.sort(reverse=True)
# # l

# # import copy

# # ll = copy.copy(l)
# # id(ll) == id(l)


# # lst3 = [1, [2, 3, 4], 5, [6, 7, 8]]
# # ll3 = copy.deepcopy(lst3)

# # type(l) is list
# # type(l) is str
# # type(l) is int
# # type(l) is float

# # a = ()
# # type(a)
# # len(a)
# # a = (
# #     1,
# #     2,
# # )
# # sorted(a)
# # a = (
# #     5,
# #     4,
# #     3,
# #     2,
# #     1,
# # )
# # b = {1, 3, 4, 5, 6}
# # type(b)
# # sorted(b)
# # b.add(2)
# # b.add(9)
# # b.add(8)
# # b
# # {2, 4, 6} <= {1, 2, 3, 4, 5, 6}
# # {2, 4, 6} >= {2, 4, 6}
# # {2, 4, 6} <= {2, 4, 6}

# # dic = {"2": "22", "3": "333"}
# # dic.keys()
# # v = list(dic.keys())
# # type(v)
# # len(dic)
# # len(v)
# # type(v)
# # type(dic)
# # d2 = (1, 2, 4)
# # type(d2)
# # v = dic.items()
# # type(v)
# # c = {"a": "aa"} | {"b": "bb"} | {"b": "ccc"}

# # for v in c:
# #     print(v)
# #     print(c[v])

# # for v in c.values():
# #     print(v)


# # x = ["a", "b", "c"]
# # y = ["d", "e", "f"]
# # z = ["g", "h", "i"]
# # vv = [x, y, x]
# # vv
# # vvv = [*x, *y, *z]
# # vvv

# # t = "qbcde"
# # tl = [*t]
# # tl

# # a = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
# # import random

# # random.shuffle(a)
# # a

# # l6 = [v for v in range(1, 44)]
# # random.choice(l6)
# # lt = random.sample(l6, 6)
# # lt.sort()
# # d = {"apple": " りんご", "orange": " みかん", "lemon": " レモン"}

# # for v in d:
# #     print(v, d[v])

# # for v in d.keys():
# #     print(v)

# # for v in d.values():
# #     print(v)

# # for v in range(97, 97 + 26):
# #     print(chr(v))

# # try:
# #     ir = iter([1, 2, 3, 4, 5])
# #     while 1:
# #         print(next(ir))
# # except StopIteration:
# #     print(v)

# # l1 = [v for v in "abcdefg"]
# # l2 = [v for v in range(5)]
# # l3 = zip(l1, l2)
# # l3

# # s = "abcd"

# # for v in enumerate(s):
# #     print(v)

# # print(s)
# # print(enumerate(s))

# # while 1:
# #     print("tt")
# # else:
# #     print("else")


# # s = "{},{},{}".format("a", "b", "c")
# # s
# # s = "{2}uuuuuuu{1}{0}".format("a", "b", "c")
# # s

# # v1 = "v1"
# # v2 = "v2"
# # v3 = "v3"
# # s = f"{v1}{v2}{v3}"
# # s

# # s = "%s,,,%s,,,%s" % ("a", "b", "c")
# # s


# # import sys

# # sys.stdout.write("abc")

# # import io

# # sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
# # y = input("in")
# # s = sys.stdin.readline()
# # sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding="utf-8")


# # # f = open("./python/rl.txt", "r",encoding='utf-8')
# # f = open("./python/rl.txt", "rb")
# # # print(f)

# # while 1:
# #     s = f.readline().decode("utf-8")
# #     if s:
# #         print(s)
# #     else:
# #         break
# # f.close()


# # a = "日本語"
# # aa = a.encode("utf-8")
# # aa
# # aaa = aa.decode("utf-8")
# # aaa


# # b = bytes([230, 150, 135])
# # b
# # b.decode("utf-8")
# # b

# # f = open("./python/rl.txt", "r", encoding="utf-8")
# # type(f)
# # for s in f:
# #     print(s.rstrip())
# # f.close()

# # f = open("./python/rl.txt", "a", encoding="utf-8")

# # f.write("f.write")
# # print("printOut", file=f)


# # f.close()

# # f = open("./python/rl.txt", "r", encoding="utf-8")
# # l = f.readlines()
# # print(l)
# # f.close()

# # coding:utf-8
# # from kivy.app import App
# # from kivy.uix.label import Label


# # class kivy01(App):
# #     def build(self):
# #         self.lb1 = Label(text="this is a test of kivy")
# #         return self.lb1


# # ap = kivy01()
# # ap.run()


# try:
#     f = open("./rl.txt", "a", encoding="utf-8")
#     fd = input("input")
#     dl = [1, 2, 3, 4, 5]
#     f.writelines(dl)
#     fd = input("input2")
#     print(fd, file=f)
#     f.close()

# except:
#     print("excpte Error")
#     f.close()

# finally:
#     f = open("./rl.txt", "r", encoding="utf-8")
#     for j in f:
#         print(j)
#     f.close()


# import os

# os.getcwd()
# os.getcwdb()
# os.listdir()
# os.getlogin()
# os.path.getsize("data.db")
# os.listdir("c:")

# os.path.exists("API")

# os.path.isfile("data.db")
# os.path.isdir("api")
# os.remove()
# os.rmdir()
# import os

# print("kkk", __file__)
# print(os.path.dirname(__file__))
# print(os.path.abspath(__file__))
# p = "./pl.txt"
# b = os.path.basename(p)
# print(os.path.dirname(p))
# print(os.path.split(p))
# os.path.splitext(p)
# os.path.abspath(p)
# os.path.isabs(os.path.abspath(p))

# from pathlib import Path

# Path.cwd()
# Path.home()
# p = Path.cwd()
# p.is_dir()
# for v in p.glob("*"):
#     if v.is_file():
#         print(v)
# p.parent

# p.as_uri()
# import sys

# print(sys.argv)


# tbl = [[x + y for y in range(1, 5)] for x in range(10, 40, 10)]

# import csv
# from pathlib import Path

# f = Path("./python/rl.txt")
# tbl = [[x for x in range(1, 5)] for y in range(10, 40, 10)]
# f.write_text(str(tbl), encoding="utf-8")

# f = Path("./python/rl.txt")
# f.read_text()


# if f.is_file():
#     f.read_text(encoding="utf-8")
# fl = list(f.parent.glob("*"))
# for f in fl:
#     print(f)

# import csv

# f = open("./python/t2.txt", "a", encoding="utf-8", newline="")
# bl = [[x + y for x in range(1, 5)] for y in range(10, 40, 10)]
# cw = csv.writer(f)

# for r in bl:
#     cw.writerow(r)
# f.close()

# f = open("./python/t2.txt", "r", encoding="utf-8")

# for v in f.readline():
#     print(v)
# f.close()

# f = open("./python/t2.txt", "r", encoding="utf-8")
# rw = csv.reader(f)

# for r in rw:
#     print(r)

# f = open("./python/t2.txt", "r", encoding="utf-8")
# cr = csv.reader(f)
# tbl = [r for r in cr]

# tbl

# tr = [x + y for x in range(1, 5) for y in range(1, 5)]
# tr = {x: y for x in range(1, 5) for y in range(2, 6)}
# tr

# import csv

# f = open("./python/t3.txt", "r", encoding="utf-8", newline="")
# ccw = csv.DictReader(f)

# d = [dict(r) for r in ccw]
# f.close()
# d


# import csv

# f = open("./python/t3.txt", "r", encoding="utf-8", newline="")
# ccw = csv.DictReader(f)

# d = [r for r in ccw]
# f.close()
# d


# def k(x, y=12):
#     z = x + y
#     return z


# a = k(12)
# print(a)


# def arg(*a):
#     print(a)
#     l = 0
#     for m in a:
#         l = l + m
#     print(l)
#     return len(a)


# def f(*a):
#     print(a)


# a = ["x", "y", "z"]
# f(a)

# a = ("a", "b", "c")
# f(a)

# a = "abcde"
# f(a)


# def arg(**a):
#     print(a["name"])
#     print(a["age"])
#     n = len(a)
#     return n


# a = arg(name="my", age=14, qq="abc")
# print(a)


# def fk2(*p, a):
#     print(p)
#     print(a)


# fk2(1, 2, 3, a=5)


# def fn(a, d, /):
#     print(a)
#     print(d)


# fn("a", "d")

# a = [4, 5, 6]
# b = [1, *a, 3]
# c = [8, 0, 22]


# def fk(*, a):
#     print(a)


# gv = 12


# def s1():
#     gv
#     print(gv)


# s1()


# def f1(x):
#     y = 2 * x

#     def f2(a):
#         nonlocal y
#         y += a
#         print(y)

#     f2(3)
#     print("y")
#     return y


# f1(5)


# from openpyxl import workbook
# import sys
# import os

# os.getcwd()
# wb =workbook()
# wb.save('./python/vvvvv.xl')
# from openpyxl import load_workbooo;


from openpyxl import Workbook  # 「pip install openpyxl」でインストールしておく

# ワークブックの新規作成と保存
wb = Workbook()
wb.save("myworkbook.xlsx")

# ワークブックの読み込み
from openpyxl import load_workbook

wb = load_workbook("myworkbook.xlsx")

# ワークシートの選択
ws = wb["Sheet"]  # ワークシートを指定
ws = wb.active  # アクティブなワークシートを選択
print(f"sheet name: {ws.title}")  # sheet name: Sheet

# ワークシートの作成
wb.create_sheet("my sheet")
wb.create_sheet("my sheet")
wb.create_sheet("my sheet33333")

# ワークシートの列挙
for sheet in wb:
    print(f"sheet name: {sheet.title}")
# 実行結果：
# Sheet
# my sheet
# my sheet1
# my sheet2

# セルに書き込み
ws["A1"] = "Hello from Python"
wb.save("myworkbook.xlsx")  # overwrite myworkbook.xlsx

# セルの値の表示
print(ws["a1"].value)  # Hello from Python

# cellメソッドでセルに書き込み
ws.cell(row=1, column=1).value = 1
print(ws["a1"].value)  # 1
ws.cell(row=1, column=1, value=2)
print(ws["a1"].value)  # 2

for idx in range(1, 4):
    ws.cell(row=1, column=idx, value=idx)
    ws.cell(row=2, column=idx, value=-idx)

# ワークシートの全セルを反復（行ごと）
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)
# 出力結果：
# 1
# 2
# 3
# -1
# -2
# -3

for row in ws.rows:
    for cell in row:
        print(cell.value)

min_row = ws.min_row
max_row = ws.max_row
min_col = ws.min_column
max_col = ws.max_column

for row in ws.iter_rows(
    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
):
    for cell in row:
        print(cell.value)

# セルの値を取得する
for row in ws.iter_rows(values_only=True):
    for value in row:
        print(value)

# ワークシートの全セルを反復（列ごと）
for column in ws.iter_cols(values_only=True):
    print(column)

for column in ws.columns:
    for cell in column:
        print(cell.value)

# セル範囲の取得
rng = ws["A1":"C2"]
print(rng)
# 出力結果：
# ((<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>),
# (<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>))

values = []
for row in rng:
    tmp = []
    for col in row:
        tmp.append(col.value)
    values.append(tmp)
# values = [[col.value for col in row] for row in rng]
print(values)

# セル範囲への代入
from openpyxl.cell.cell import Cell


def get_shape(rng):
    return (len(rng), len(rng[0]))


def assign2range(dst, src):
    dst_shape = get_shape(dst)
    src_shape = get_shape(src)
    if src_shape != dst_shape:
        raise ValueError("shapes of arguments not match")

    for d, s in zip(dst, src):  # dst／src: Cellまたは値のタプルを格納するタプル
        for t, v in zip(d, s):  # d／s: Cellまたは値を格納するタプル
            if isinstance(v, Cell):
                v = v.value
            t.value = v


values = [[1, 2, 3], [4, 5, 6]]

assign2range(ws["A1":"C2"], values)
assign2range(ws["A1:C1"], [[100, 200, 300]])
assign2range(ws["A3":"C4"], ws["A1":"C2"])

wb.save("myworkbook.xlsx")
